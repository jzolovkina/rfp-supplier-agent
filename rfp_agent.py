"""
RFP Supplier Agent — Streamlit app · Zinit brand
Deploy on share.streamlit.io
"""

import json, time, io, re
import streamlit as st
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Zinit · Supplier Search",
    page_icon="🔍",
    layout="wide",
)

# ─── Zinit brand CSS ────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rubik:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"], .stApp, div, p, span, label, input, button {
    font-family: 'Rubik', sans-serif !important;
}

.stApp { background: #F4F2FF; }

/* hide chrome */
#MainMenu, footer { visibility: hidden; }
.block-container { padding-top: 0 !important; padding-bottom: 40px !important; max-width: 1160px; }

/* ─ Topbar ─ */
.zn-topbar {
    background: #3D0099;
    padding: 0 40px;
    height: 56px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin: -1rem -1rem 0;
}
.zn-logo-svg { height: 20px; display: block; }
.zn-topbar-right {
    display: flex;
    align-items: center;
    gap: 10px;
}
.zn-badge {
    background: rgba(255,255,255,.12);
    border: 1px solid rgba(255,255,255,.2);
    color: rgba(255,255,255,.85);
    font-size: 11px;
    font-weight: 500;
    padding: 3px 12px;
    border-radius: 20px;
    letter-spacing: .3px;
}

/* ─ Hero ─ */
.zn-hero {
    background: linear-gradient(135deg, #1A0050 0%, #3D0099 55%, #6B21D4 100%);
    padding: 36px 48px 40px;
    margin: 0 -1rem 28px;
    position: relative;
    overflow: hidden;
}
.zn-hero::before {
    content:'';position:absolute;top:-60px;right:-60px;
    width:320px;height:320px;border-radius:50%;
    background:radial-gradient(circle,rgba(109,70,229,.45) 0%,transparent 70%);
}
.zn-hero::after {
    content:'';position:absolute;bottom:-80px;left:35%;
    width:240px;height:240px;border-radius:50%;
    background:radial-gradient(circle,rgba(6,182,212,.22) 0%,transparent 70%);
}
.zn-hero-label {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: .14em;
    text-transform: uppercase;
    color: #7EDCE8;
    margin-bottom: 10px;
    position: relative;
    z-index: 1;
}
.zn-hero-title {
    font-size: 30px;
    font-weight: 700;
    color: #fff;
    line-height: 1.25;
    margin-bottom: 10px;
    position: relative;
    z-index: 1;
}
.zn-hero-sub {
    font-size: 14px;
    color: rgba(255,255,255,.62);
    line-height: 1.65;
    max-width: 560px;
    position: relative;
    z-index: 1;
    margin-bottom: 20px;
}
.zn-pills {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    position: relative;
    z-index: 1;
}
.zn-pill {
    background: rgba(255,255,255,.11);
    border: 1px solid rgba(255,255,255,.18);
    color: rgba(255,255,255,.88);
    font-size: 12px;
    font-weight: 500;
    padding: 5px 14px;
    border-radius: 20px;
}
.zn-pill-teal {
    background: rgba(6,182,212,.2);
    border-color: rgba(6,182,212,.45);
    color: #5EE3F3;
}

/* ─ Cards ─ */
.zn-card {
    background: #fff;
    border: 1px solid #E5DFFB;
    border-radius: 16px;
    padding: 28px 32px;
    margin-bottom: 20px;
    box-shadow: 0 2px 16px rgba(61,0,153,.06);
}
.zn-card-head {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 22px;
    padding-bottom: 18px;
    border-bottom: 1px solid #F0EBFA;
}
.zn-step {
    background: linear-gradient(135deg,#5B21B6,#7C3AED);
    color: #fff;
    font-size: 11px;
    font-weight: 700;
    padding: 3px 10px;
    border-radius: 8px;
    letter-spacing: .5px;
}
.zn-card-title {
    font-size: 15px;
    font-weight: 600;
    color: #1A0050;
}

/* ─ How-it-works strip ─ */
.zn-how {
    display: grid;
    grid-template-columns: repeat(2,1fr);
    gap: 12px;
    margin-bottom: 0;
}
.zn-how-item {
    background: #F8F5FF;
    border: 1px solid #E5DFFB;
    border-radius: 12px;
    padding: 16px 18px;
}
.zn-how-label {
    font-size: 11px;
    font-weight: 700;
    color: #7C3AED;
    text-transform: uppercase;
    letter-spacing: .1em;
    margin-bottom: 5px;
}
.zn-how-label-teal { color: #0891B2; }
.zn-how-desc { font-size: 13px; color: #4A3575; line-height: 1.5; }

/* ─ Status chip ─ */
.zn-status {
    display: inline-block;
    background: rgba(6,182,212,.1);
    color: #0891B2;
    border: 1px solid rgba(6,182,212,.3);
    font-size: 11px;
    font-weight: 600;
    padding: 2px 10px;
    border-radius: 20px;
    text-transform: uppercase;
    letter-spacing: .5px;
}

/* ─ Streamlit overrides ─ */
.stTextInput input, .stNumberInput input {
    border: 1.5px solid #DDD6F0 !important;
    border-radius: 10px !important;
    font-family: 'Rubik',sans-serif !important;
    font-size: 13px !important;
    color: #1A0050 !important;
    background: #FAFAFE !important;
    padding: 10px 14px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: #7C3AED !important;
    box-shadow: 0 0 0 3px rgba(124,58,237,.12) !important;
}

/* Primary → purple gradient */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg,#5B21B6 0%,#7C3AED 100%) !important;
    border: none !important;
    border-radius: 10px !important;
    color: #fff !important;
    font-family: 'Rubik',sans-serif !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    padding: 11px 32px !important;
    letter-spacing: .2px !important;
    box-shadow: 0 4px 14px rgba(91,33,182,.35) !important;
    transition: all .2s !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 22px rgba(91,33,182,.45) !important;
}
.stButton > button[kind="primary"]:disabled {
    background: #C4B8E8 !important;
    box-shadow: none !important;
    transform: none !important;
}

/* Download → teal */
.stDownloadButton > button {
    background: linear-gradient(135deg,#0E7490 0%,#06B6D4 100%) !important;
    border: none !important;
    border-radius: 10px !important;
    color: #fff !important;
    font-family: 'Rubik',sans-serif !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    padding: 11px 32px !important;
    box-shadow: 0 4px 14px rgba(6,182,212,.35) !important;
    transition: all .2s !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 22px rgba(6,182,212,.45) !important;
}

/* File uploader */
[data-testid="stFileUploader"] {
    background: #FAFAFE !important;
    border: 2px dashed #C4B8E8 !important;
    border-radius: 14px !important;
    padding: 8px !important;
}
[data-testid="stFileUploader"]:hover { border-color: #7C3AED !important; }

/* Progress bar */
.stProgress > div > div {
    background: linear-gradient(90deg,#5B21B6,#06B6D4) !important;
    border-radius: 4px !important;
}
/* Dataframe */
[data-testid="stDataFrame"] {
    border: 1px solid #E5DFFB !important;
    border-radius: 12px !important;
    overflow: hidden !important;
}
/* Metrics */
[data-testid="stMetric"] {
    background: #FAFAFE !important;
    border: 1px solid #E5DFFB !important;
    border-radius: 12px !important;
    padding: 16px 20px !important;
}
[data-testid="stMetricValue"] { color: #1A0050 !important; font-weight: 700 !important; }
[data-testid="stMetricDelta"] { color: #06B6D4 !important; }

/* Alerts */
.stSuccess { background: rgba(6,182,212,.07) !important; border-color: rgba(6,182,212,.3) !important; border-radius: 10px !important; }
.stInfo    { background: rgba(91,33,182,.05) !important; border-color: rgba(91,33,182,.2) !important; border-radius: 10px !important; }
.stError   { border-radius: 10px !important; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: #1A0050 !important;
    border-right: 1px solid rgba(255,255,255,.07) !important;
}
[data-testid="stSidebar"] * { color: rgba(255,255,255,.82) !important; }
[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,[data-testid="stSidebar"] h3 { color: #fff !important; }
[data-testid="stSidebar"] .stTextInput input,
[data-testid="stSidebar"] .stNumberInput input {
    background: rgba(255,255,255,.08) !important;
    border-color: rgba(255,255,255,.15) !important;
    color: #fff !important;
}
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,.1) !important; }
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p { color: rgba(255,255,255,.65) !important; }

hr { border-color: #EDE8FA !important; }
.stSpinner > div { border-top-color: #7C3AED !important; }
</style>

<!-- topbar -->
<div class="zn-topbar">
  <svg class="zn-logo-svg" viewBox="0 0 78 24" fill="none" xmlns="http://www.w3.org/2000/svg">
    <path d="M10.1219 18.2687H15.5444V24H0L5.42245 14.3284H0V8.59701H15.5444L10.1219 18.2687Z" fill="white" fill-opacity=".87"/>
    <path d="M26.0278 24H20.2438V8.59701H26.0278V24Z" fill="white" fill-opacity=".87"/>
    <path d="M38.5968 8.59701C40.0676 8.59701 41.3447 8.59462 42.394 8.67957C43.4769 8.76726 44.5693 8.96209 45.6249 9.49498C47.1893 10.2849 48.4614 11.5454 49.2585 13.0956C49.7963 14.1415 49.9929 15.224 50.0814 16.2971C50.1672 17.3368 50.1647 18.6024 50.1647 20.0597V24H44.3808V20.0597C44.3808 18.5079 44.3786 17.5163 44.3165 16.7641C44.2573 16.0456 44.1583 15.801 44.1051 15.6975C43.8625 15.2257 43.4752 14.842 42.9991 14.6016C42.8947 14.5489 42.6478 14.4507 41.9227 14.392C41.1636 14.3306 40.1629 14.3284 38.5968 14.3284H36.5112V24H30.7272V8.59701H38.5968Z" fill="white" fill-opacity=".87"/>
    <path d="M60.7315 24H54.9475V8.59701H60.7315V24Z" fill="white" fill-opacity=".87"/>
    <path d="M71.1316 8.59701H78V13.9701H71.1362C71.1417 14.7517 71.1554 15.343 71.1958 15.8329C71.2551 16.5514 71.3541 16.7961 71.4073 16.8995C71.6499 17.3713 72.0372 17.755 72.5133 17.9954C72.6177 18.0481 72.8645 18.1463 73.5897 18.205C74.3488 18.2664 75.3494 18.2687 76.9155 18.2687H78V24H76.9155C75.4448 24 74.1677 24.0024 73.1184 23.9174C72.0354 23.8298 70.943 23.6349 69.8875 23.102C68.323 22.3121 67.051 21.0516 66.2538 19.5014C65.716 18.4555 65.5194 17.373 65.4309 16.2999C65.3747 15.6174 65.3567 14.8376 65.3508 13.9701H65.3476V3.22388H71.1316V8.59701Z" fill="white" fill-opacity=".87"/>
    <path d="M23.1358 0C24.9326 0 26.3893 1.44338 26.3893 3.22388C26.3893 5.00438 24.9326 6.44776 23.1358 6.44776C21.339 6.44776 19.8823 5.00438 19.8823 3.22388C19.8823 1.44338 21.339 0 23.1358 0Z" fill="white" fill-opacity=".87"/>
    <path d="M57.8395 0C59.6363 0 61.093 1.44338 61.093 3.22388C61.093 5.00438 59.6363 6.44776 57.8395 6.44776C56.0427 6.44776 54.586 5.00438 54.586 3.22388C54.586 1.44338 56.0427 0 57.8395 0Z" fill="white" fill-opacity=".87"/>
  </svg>
  <div class="zn-topbar-right">
    <span class="zn-badge">Supplier Search</span>
    <span class="zn-badge">AI-powered</span>
  </div>
</div>

<!-- hero -->
<div class="zn-hero">
  <div class="zn-hero-label">SSM Tool · Zinit Procurement Platform</div>
  <div class="zn-hero-title">Supplier & Contact Search</div>
  <div class="zn-hero-sub">
    Upload a list of active RFPs — the agent finds verified supplier companies
    and locates the right commercial contact at each one for SSM outreach.
  </div>
  <div class="zn-pills">
    <span class="zn-pill">Stage 1 — Company discovery</span>
    <span class="zn-pill zn-pill-teal">Stage 2 — Persona contact match</span>
    <span class="zn-pill">Excel export ready for SSM</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:8px 0 20px;">
      <svg viewBox="0 0 78 24" fill="none" xmlns="http://www.w3.org/2000/svg" style="height:18px;display:block;margin-bottom:16px;">
        <path d="M10.1219 18.2687H15.5444V24H0L5.42245 14.3284H0V8.59701H15.5444L10.1219 18.2687Z" fill="white" fill-opacity=".87"/>
        <path d="M26.0278 24H20.2438V8.59701H26.0278V24Z" fill="white" fill-opacity=".87"/>
        <path d="M38.5968 8.59701C40.0676 8.59701 41.3447 8.59462 42.394 8.67957C43.4769 8.76726 44.5693 8.96209 45.6249 9.49498C47.1893 10.2849 48.4614 11.5454 49.2585 13.0956C49.7963 14.1415 49.9929 15.224 50.0814 16.2971C50.1672 17.3368 50.1647 18.6024 50.1647 20.0597V24H44.3808V20.0597C44.3808 18.5079 44.3786 17.5163 44.3165 16.7641C44.2573 16.0456 44.1583 15.801 44.1051 15.6975C43.8625 15.2257 43.4752 14.842 42.9991 14.6016C42.8947 14.5489 42.6478 14.4507 41.9227 14.392C41.1636 14.3306 40.1629 14.3284 38.5968 14.3284H36.5112V24H30.7272V8.59701H38.5968Z" fill="white" fill-opacity=".87"/>
        <path d="M60.7315 24H54.9475V8.59701H60.7315V24Z" fill="white" fill-opacity=".87"/>
        <path d="M71.1316 8.59701H78V13.9701H71.1362C71.1417 14.7517 71.1554 15.343 71.1958 15.8329C71.2551 16.5514 71.3541 16.7961 71.4073 16.8995C71.6499 17.3713 72.0372 17.755 72.5133 17.9954C72.6177 18.0481 72.8645 18.1463 73.5897 18.205C74.3488 18.2664 75.3494 18.2687 76.9155 18.2687H78V24H76.9155C75.4448 24 74.1677 24.0024 73.1184 23.9174C72.0354 23.8298 70.943 23.6349 69.8875 23.102C68.323 22.3121 67.051 21.0516 66.2538 19.5014C65.716 18.4555 65.5194 17.373 65.4309 16.2999C65.3747 15.6174 65.3567 14.8376 65.3508 13.9701H65.3476V3.22388H71.1316V8.59701Z" fill="white" fill-opacity=".87"/>
        <path d="M23.1358 0C24.9326 0 26.3893 1.44338 26.3893 3.22388C26.3893 5.00438 24.9326 6.44776 23.1358 6.44776C21.339 6.44776 19.8823 5.00438 19.8823 3.22388C19.8823 1.44338 21.339 0 23.1358 0Z" fill="white" fill-opacity=".87"/>
        <path d="M57.8395 0C59.6363 0 61.093 1.44338 61.093 3.22388C61.093 5.00438 59.6363 6.44776 57.8395 6.44776C56.0427 6.44776 54.586 5.00438 54.586 3.22388C54.586 1.44338 56.0427 0 57.8395 0Z" fill="white" fill-opacity=".87"/>
      </svg>
    </div>
    """, unsafe_allow_html=True)

    st.header("Settings")
    api_key = st.text_input(
        "Anthropic API Key",
        type="password",
        placeholder="sk-ant-api03-...",
    )
    st.caption("Stored in session only — never saved to disk.")

    supplier_count_override = st.number_input(
        "Suppliers per RFP (override)",
        min_value=0, max_value=50, value=0,
        help="Set to 0 to use value from the Excel file. Start with 3–5 to test quality first.",
    )
    st.divider()
    st.markdown("**How it works:**")
    st.markdown("**Stage 1** — Finds N verified companies via web search")
    st.markdown("**Stage 2** — Finds the right commercial contact at each company")
    st.markdown("*(Sales Manager, KAM, BD Manager, Tender Manager...)*")
    st.divider()
    st.markdown("**Tip:** Start with 3–5 suppliers to verify quality before running 20.")


def normalize_header(text):
    s = str(text or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def parse_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    if "RFP_Input" not in wb.sheetnames:
        st.error('Sheet "RFP_Input" not found.')
        return [], []

    ws = wb["RFP_Input"]
    rows_raw = list(ws.iter_rows(values_only=True))

    header_row_idx = -1
    for i, row in enumerate(rows_raw):
        if normalize_header(row[0]) == "rfp_id":
            header_row_idx = i
            break

    if header_row_idx < 0:
        st.error('Header "RFP ID" not found in RFP_Input sheet.')
        return [], []

    headers = [normalize_header(c) for c in rows_raw[header_row_idx]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, *keys):
        for k in keys:
            if k in idx:
                v = str(row[idx[k]] or "").strip()
                if v:
                    return v
        return ""

    rfp_rows = []
    for row in rows_raw[header_row_idx + 2:]:
        if not row or row[0] is None:
            continue
        rid = str(row[idx.get("rfp_id", 0)] or "").strip()
        if not rid or re.match(r"^(unique|e\.g\.|links|\s*$)", rid, re.I):
            continue
        sc = get(row, "supplier_count")
        try:
            supplier_count = int(float(sc)) if sc else 20
        except ValueError:
            supplier_count = 20

        rfp_rows.append({
            "rfp_id":         rid,
            "title":          get(row, "rfp_title", "title"),
            "description":    get(row, "description"),
            "country":        get(row, "country"),
            "region":         get(row, "region_city", "region", "city"),
            "phase":          get(row, "rfp_phase", "phase"),
            "keywords":       get(row, "keywords_tags", "keywords", "tags"),
            "supplier_count": supplier_count,
            "notes":          get(row, "notes"),
            "output_sheet":   get(row, "output_sheet") or ("Results_" + rid),
        })

    spec_rows = []
    if "SPEC_Items" in wb.sheetnames:
        ws2 = wb["SPEC_Items"]
        rows2 = list(ws2.iter_rows(values_only=True))
        h2_idx = -1
        for i, row in enumerate(rows2):
            if normalize_header(row[0]) == "rfp_id":
                h2_idx = i
                break
        if h2_idx >= 0:
            heads2 = [normalize_header(c) for c in rows2[h2_idx]]
            i2 = {h: i for i, h in enumerate(heads2)}
            for row in rows2[h2_idx + 2:]:
                if not row or row[0] is None:
                    continue
                sid = str(row[i2.get("rfp_id", 0)] or "").strip()
                if not sid or re.match(r"links to|rfp_id", sid, re.I):
                    continue
                spec_rows.append({
                    "rfp_id":      sid,
                    "item_name":   str(row[i2.get("item_name", 1)] or "").strip(),
                    "description": str(row[i2.get("item_description", 3)] or "").strip(),
                    "quantity":    str(row[i2.get("quantity", 4)] or "").strip(),
                    "unit":        str(row[i2.get("unit_of_measure", 5)] or "").strip(),
                })

    return rfp_rows, spec_rows


def call_claude(client, prompt, max_tokens, single_object):
    messages = [{"role": "user", "content": prompt}]
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=max_tokens,
        tools=[{"type": "web_search_20250305", "name": "web_search"}],
        messages=messages,
    )
    full_text = "".join(b.text for b in response.content if hasattr(b, "text"))

    if not full_text.strip():
        tool_results = [
            {"type": "tool_result", "tool_use_id": b.id, "content": "Search completed."}
            for b in response.content if b.type == "tool_use"
        ]
        if tool_results:
            follow_up = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=max_tokens,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=messages + [
                    {"role": "assistant", "content": response.content},
                    {"role": "user", "content": tool_results},
                ],
            )
            full_text = "".join(b.text for b in follow_up.content if hasattr(b, "text"))

    if not full_text.strip():
        raise ValueError("Agent returned empty response.")

    if single_object:
        m = re.search(r"\{[\s\S]*\}", full_text)
        if not m:
            raise ValueError(f"No JSON object found. Preview: {full_text[:200]}")
        return json.loads(m.group())
    else:
        m = re.search(r"\[[\s\S]*\]", full_text)
        if not m:
            raise ValueError(f"No JSON array found. Preview: {full_text[:200]}")
        result = json.loads(m.group())
        if not isinstance(result, list):
            raise ValueError("Response is not a list.")
        return result


def find_companies(client, rfp, specs, count):
    has_spec = len(specs) > 0
    spec_text = ""
    if has_spec:
        lines = []
        for i, s in enumerate(specs, 1):
            line = f"  {i}. {s['item_name']}"
            if s["description"]:
                line += f" — {s['description']}"
            if s["quantity"]:
                line += f" | Qty: {s['quantity']} {s['unit']}"
            lines.append(line)
        spec_text = "\n\nSPECIFICATION (exact items/works needed):\n" + "\n".join(lines)

    # Detect tender type from keywords + description
    combined = (rfp["keywords"] + " " + rfp["description"]).lower()
    service_words = ["jasa", "service", "maintenance", "rental", "sewa",
                     "installation", "repair", "pemasangan", "perawatan",
                     "konstruksi", "construction", "works", "contractor"]
    is_service = any(w in combined for w in service_words)
    tender_type = "services / works" if is_service else "goods / products"

    location = rfp["region"] or rfp["country"]

    prompt = f"""You are a senior B2B procurement specialist. Your task is to find {count} REAL, VERIFIABLE companies for this tender through web search.

TENDER:
- ID: {rfp['rfp_id']}
- Title: {rfp['title']}
- Type: {tender_type}
- Country: {rfp['country']}
- Delivery / service location: {location}
- Keywords: {rfp['keywords']}
- Description: {rfp['description'][:700]}{spec_text}
{('- Additional notes: ' + rfp['notes']) if rfp['notes'] else ''}

MANDATORY SEARCH CRITERIA — every result must meet ALL of these:

1. LEGAL ENTITY
   Only formal registered companies (PT, CV, or local equivalent). No individuals, no unregistered vendors.

2. LOCATION
   Must be able to deliver goods or provide services in: {location}.
   Strongly prefer companies based in or near {location}.

3. SUPPLY CHAIN ROLE
   For goods: find suppliers, distributors, or authorized/official dealers — NOT online marketplaces.
   For services: find contractors or service providers with relevant experience.
   EXCLUDE: Tokopedia, Shopee, Lazada, Bukalapak, Amazon, eBay, Alibaba, and any other e-commerce marketplace.

4. B2B POSTURE
   Company must actively sell to businesses and participate in procurement or tenders.
   Avoid consumer-only brands or retail-only stores.

5. ACTIVE & VERIFIABLE
   Must have a real website, Google Maps listing, or verifiable industry directory entry.
   Only include companies you actually found through web search — not guessed or invented.

6. SPECIFIC FIT
   {"Prefer authorized distributors or official dealers for any branded products mentioned." if not is_service else "Must have proven experience in this specific type of service. Look for KBLI codes, project portfolio, or client references."}
   {"Confirm they can supply the specific items or brands in the specification." if has_spec else "Confirm their core business matches what is being tendered."}

ANTI-HALLUCINATION RULES:
- If you cannot find {count} verified companies, return FEWER real ones. Quality over quantity.
- Never invent a company name or guess a website URL.
- Set "verified": false if you are not 100% certain the company exists and matches.

Return ONLY a valid JSON array. No markdown, no explanation, nothing outside the brackets.

[{{
  "company_name": "PT Example Supplier",
  "website": "https://example.co.id",
  "country": "{rfp['country']}",
  "city": "Jakarta",
  "legal_type": "PT",
  "supply_chain_role": "Authorized distributor / Contractor / Supplier",
  "verified": true,
  "company_description": "2-3 sentences: what they supply/do and who their clients are",
  "why_relevant": "Specific reason why they are a strong match for THIS tender",
  "relevance_score": 8
}}]"""

    return call_claude(client, prompt, 6000, False)


def find_contact(client, company, rfp):
    prompt = f"""Find a commercial contact person at this company who is responsible for B2B sales and tender participation.

COMPANY: {company['company_name']}
WEBSITE: {company.get('website', 'unknown')}
COUNTRY: {company.get('country', rfp['country'])}
CONTEXT: This company will be invited to a procurement tender for: {rfp['title']} ({rfp['keywords']})

TARGET PERSONA — who we are looking for:
A person responsible for: attracting new clients, participating in tenders/RFPs, B2B sales.
Motivated by: won tenders, contract volume, new corporate clients.

TARGET JOB TITLES (in priority order):
Sales Manager, B2B Sales Manager, Corporate Sales Manager, Enterprise Sales Manager,
Key Account Manager (KAM), Business Development Manager, Head of Business Development,
Head of Sales, Sales Director, Commercial Director,
Tender Manager, Bid Manager, Proposal Manager, RFP Manager, Pre-Sales Manager,
Partnership Manager, Strategic Partnerships Manager.

NOT looking for: secretary, office manager, HR, customer support, accountant, finance.

CONTACT PRIORITY (search in this order):
1. Direct mobile phone number
2. Personal email (NOT info@, contact@, sales@, or other generic inboxes)
3. LinkedIn profile URL
4. General company email as last resort only

Search LinkedIn, the company website, industry directories, and other sources.
If you find a matching person → set "contact_found": true.
If no matching person found → set "contact_found": false and provide whatever general company contacts you can find.

Return ONLY a JSON object. No markdown, no text outside the braces.
{{
  "contact_found": true,
  "contact_person": "First Last",
  "job_title": "Sales Manager",
  "phone": "direct mobile or empty string",
  "email": "direct personal email or empty string",
  "linkedin": "https://linkedin.com/in/... or empty string",
  "general_email": "info@company.com or empty string",
  "general_phone": "+xx xxx xxx xxxx or empty string",
  "contact_note": "where found, or why persona was not found"
}}"""

    return call_claude(client, prompt, 2000, True)


THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MID_BLUE = "2E5FA3"
DARK_BLUE = "1F3864"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
GREY_BG = "F2F2F2"


def build_output_excel(original_bytes, rfp_rows, results):
    # Load the original file as-is — RFP_Input and SPEC_Items keep 100% of their formatting
    wb_out = openpyxl.load_workbook(io.BytesIO(original_bytes))

    # Remove all old Results_ sheets and Instructions from the output workbook
    for sheet_name in list(wb_out.sheetnames):
        if sheet_name.startswith("Results_") or sheet_name == "📋 Instructions":
            del wb_out[sheet_name]

    HEADERS = [
        "#", "Company Name", "Website", "Country", "City",
        "Legal Type", "Supply Chain Role", "Verified?",
        "Contact Person", "Job Title",
        "Direct Phone", "Direct Email", "LinkedIn",
        "General Email", "General Phone",
        "Company Description", "Why Relevant", "Relevance Score (1-10)",
        "Persona Match?", "Contact Notes",
    ]
    COL_WIDTHS = [
        4, 26, 26, 13, 15,
        10, 22, 10,
        22, 24,
        18, 28, 28,
        24, 18,
        36, 36, 12,
        18, 30,
    ]

    for rfp in rfp_rows:
        suppliers = results.get(rfp["rfp_id"], [])
        sheet_name = rfp["output_sheet"][:31]
        ws = wb_out.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = "27AE60"
        ws.sheet_view.showGridLines = False

        ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
        banner = ws["A1"]
        banner.value = (
            f"SUPPLIER LIST FOR SSM  |  ID {rfp['rfp_id']}: {rfp['title']}"
            f"  |  {rfp['country']}"
            + (f"  |  Phase: {rfp['phase']}" if rfp["phase"] else "")
        )
        banner.font = Font(name="Arial", bold=True, color=WHITE, size=12)
        banner.fill = PatternFill("solid", start_color=DARK_BLUE)
        banner.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{get_column_letter(len(HEADERS))}2")
        ctx = ws["A2"]
        with_contact = sum(1 for s in suppliers if s.get("contact_found"))
        verified = sum(1 for s in suppliers if s.get("verified", True))
        ctx.value = (
            f"Keywords: {rfp['keywords']}  |  "
            f"Companies found: {len(suppliers)}  |  "
            f"Verified: {verified}  |  "
            f"With persona contact: {with_contact}"
        )
        ctx.font = Font(name="Arial", italic=True, color="1F3864", size=9)
        ctx.fill = PatternFill("solid", start_color=LIGHT_BLUE)
        ctx.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", start_color=MID_BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci - 1]
        ws.row_dimensions[3].height = 30

        if not suppliers:
            ws.cell(row=4, column=1).value = "No suppliers found — check the log for errors."
        else:
            for ri, s in enumerate(suppliers):
                r = 4 + ri
                bg = WHITE if ri % 2 == 0 else GREY_BG
                row_data = [
                    ri + 1,
                    s.get("company_name", ""),
                    s.get("website", ""),
                    s.get("country", ""),
                    s.get("city", ""),
                    s.get("legal_type", ""),
                    s.get("supply_chain_role", ""),
                    "Yes" if s.get("verified", True) else "Unverified",
                    s.get("contact_person", ""),
                    s.get("job_title", ""),
                    s.get("phone", ""),
                    s.get("email", ""),
                    s.get("linkedin", ""),
                    s.get("general_email", ""),
                    s.get("general_phone", ""),
                    s.get("company_description", ""),
                    s.get("why_relevant", ""),
                    s.get("relevance_score", ""),
                    "Yes" if s.get("contact_found") else "No (general contacts)",
                    s.get("contact_note", ""),
                ]
                for ci, val in enumerate(row_data, 1):
                    c = ws.cell(row=r, column=ci, value=val)
                    c.font = Font(name="Arial", size=10)
                    c.fill = PatternFill("solid", start_color=bg)
                    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    c.border = BORDER
                ws.row_dimensions[r].height = 40

        ws.freeze_panes = "B4"

    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()


# ── Main UI ──────────────────────────────────────────────────────────────────

# ── Section 01: Upload ──
st.markdown("""
<div class="zn-card">
  <div class="zn-card-head">
    <span class="zn-step">01</span>
    <span class="zn-card-title">Upload RFP Excel File</span>
  </div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Drop your RFP_Supplier_Template.xlsx here",
    type=["xlsx", "xls"],
    label_visibility="collapsed",
)

rfp_rows, spec_rows, original_bytes = [], [], None

if uploaded:
    original_bytes = uploaded.read()
    rfp_rows, spec_rows = parse_excel(original_bytes)
    if rfp_rows:
        st.success(f"✓ Found **{len(rfp_rows)} RFP(s)** in Collecting Bids · {len(spec_rows)} specification lines")
        import pandas as pd
        preview_data = []
        for r in rfp_rows:
            preview_data.append({
                "RFP ID":    r["rfp_id"],
                "Phase":     r["phase"] or "—",
                "Title":     r["title"][:55] + ("…" if len(r["title"]) > 55 else ""),
                "Country":   r["country"],
                "Region":    r["region"] or "—",
                "Keywords":  r["keywords"][:45] + ("…" if len(r["keywords"]) > 45 else ""),
                "Suppliers": r["supplier_count"],
            })
        st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

st.markdown("</div>", unsafe_allow_html=True)

# ── Section 02: Run ──
st.markdown("""
<div class="zn-card">
  <div class="zn-card-head">
    <span class="zn-step">02</span>
    <span class="zn-card-title">Run Supplier Search</span>
  </div>
  <div class="zn-how">
    <div class="zn-how-item">
      <div class="zn-how-label">Stage 1 — Company discovery</div>
      <div class="zn-how-desc">
        Finds N verified supplier companies per RFP via web search.
        Filters: PT/CV only · local/regional · B2B · no marketplaces.
      </div>
    </div>
    <div class="zn-how-item">
      <div class="zn-how-label zn-how-label-teal">Stage 2 — Persona contact</div>
      <div class="zn-how-desc">
        For each company finds: Sales Manager · KAM · Tender Manager · BD Director.
        Fallback to general contacts if persona not found.
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

can_run = bool(rfp_rows) and bool(api_key) and api_key.startswith("sk-")

if not api_key:
    st.info("👈 Enter your Anthropic API key in the sidebar to get started.")
elif not api_key.startswith("sk-"):
    st.error("API key must start with sk-ant-...")
elif not rfp_rows:
    st.info("Upload an RFP Excel file above to continue.")

if st.button("Find Suppliers", type="primary", disabled=not can_run):
    client = anthropic.Anthropic(api_key=api_key)
    results = {}
    overall_bar = st.progress(0, text="Starting...")
    log_area = st.empty()
    log_lines = []

    def add_log(msg):
        log_lines.append(msg)
        log_area.markdown(
            "<div style='"
            "background:#0D0621;border:1px solid rgba(124,58,237,.25);border-radius:12px;"
            "padding:14px 20px;font-family:monospace;font-size:12px;max-height:320px;"
            "overflow-y:auto;line-height:1.9;color:#A89DC8;'>"
            + "<br>".join(log_lines[-40:])
            + "</div>",
            unsafe_allow_html=True,
        )

    total = len(rfp_rows)
    for i, rfp in enumerate(rfp_rows):
        count = supplier_count_override or rfp["supplier_count"] or 20
        specs = [s for s in spec_rows if s["rfp_id"] == rfp["rfp_id"]]
        overall_bar.progress(i / total, text=f"Processing {rfp['rfp_id']} ({i+1}/{total})")

        add_log(f"[{i+1}/{total}] {rfp['rfp_id']} — {rfp['title'] or '(no title)'}")
        add_log(f"  Location: {rfp['region'] or rfp['country']} | Keywords: {rfp['keywords'] or '(empty)'}")
        add_log(f"  Spec lines: {len(specs)}" if specs else "  No specification — using title & description")

        add_log(f"  [Stage 1] Searching {count} verified companies...")
        companies = []
        try:
            companies = find_companies(client, rfp, specs, count)
            verified = sum(1 for c in companies if c.get("verified", True))
            add_log(f"  [Stage 1] ✓ {len(companies)} found ({verified} verified)")
        except Exception as e:
            add_log(f"  [Stage 1] ✗ {e}")
            results[rfp["rfp_id"]] = []
            if i < total - 1:
                time.sleep(5)
            continue

        add_log(f"  [Stage 2] Finding contacts for {len(companies)} companies...")
        enriched = []
        for j, co in enumerate(companies):
            try:
                contact = find_contact(client, co, rfp)
                enriched.append({**co, **contact})
                if contact.get("contact_found"):
                    add_log(f"    ✓ {co['company_name']}: {contact.get('contact_person')} ({contact.get('job_title')})")
                else:
                    add_log(f"    ~ {co['company_name']}: no persona match — general contacts saved")
            except Exception as e:
                add_log(f"    ✗ {co['company_name']}: {e}")
                enriched.append({
                    **co,
                    "contact_found": False, "contact_person": "", "job_title": "",
                    "email": "", "phone": "", "linkedin": "",
                    "general_email": "", "general_phone": "", "contact_note": str(e),
                })
            if j < len(companies) - 1:
                time.sleep(3)

        results[rfp["rfp_id"]] = enriched
        with_contact = sum(1 for s in enriched if s.get("contact_found"))
        add_log(f"  [Stage 2] ✓ Persona contacts: {with_contact}/{len(enriched)}")
        if i < total - 1:
            add_log("  Pausing 5s before next RFP...")
            time.sleep(5)

    overall_bar.progress(1.0, text="✓ Done!")
    add_log("=" * 52)
    add_log("All RFPs processed. Download the Excel file below.")

    # ── Results ──
    st.markdown("""
<div class="zn-card" style="margin-top:20px">
  <div class="zn-card-head">
    <span class="zn-step" style="background:linear-gradient(135deg,#0E7490,#06B6D4)">✓</span>
    <span class="zn-card-title">Results — Ready for SSM</span>
  </div>
""", unsafe_allow_html=True)

    cols = st.columns(min(len(rfp_rows), 4))
    for i, rfp in enumerate(rfp_rows):
        sup = results.get(rfp["rfp_id"], [])
        wc = sum(1 for s in sup if s.get("contact_found"))
        with cols[i % len(cols)]:
            st.metric(
                label=rfp["rfp_id"],
                value=f"{len(sup)} companies",
                delta=f"{wc} with persona contact",
            )

    st.markdown("</div>", unsafe_allow_html=True)

    with st.spinner("Building Excel..."):
        out_bytes = build_output_excel(original_bytes, rfp_rows, results)

    from datetime import date
    st.download_button(
        label="⬇ Download Excel for SSM",
        data=out_bytes,
        file_name=f"RFP_Suppliers_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
